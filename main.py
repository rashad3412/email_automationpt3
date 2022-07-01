import openpyxl
import smtplib
import sys
wb = openpyxl.load_workbook('duesRecords.xlsx')
name = wb.sheetnames
sheet = wb["Sheet1"]
lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value

# TODO: Check each member's payment status.
# TODO:  Log in to email account.
# TODO: Send out reminder emails.

# Check each member's payment status
unpaidMembers = {}
for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email

# Log in to email account
smtpdObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpdObj.ehlo()
smtpdObj.starttls()
smtpdObj.login('meeztech12@gmail.com', "kbnkwjdkeiozbygr")

for name, email in unpaidMembers.items():
    body = "Subject: %s dues unpaid.\nDear %s, \nRecords show that you have not paid dues for %s.Please make this' \
           'payment as soon as possible. Thank you!'" % (latestMonth, name, latestMonth)
    print('sending email to %s...' % email)
    sendmailStatus = smtpdObj.sendmail('meeztech12@gmail.com', email, body)

    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s' % (email, sendmailStatus))
smtpdObj.quit()
